from pyimpBB.helper import *
from pyimpBB.bounding import *
from pyimpBB.solver import *
from pyimpBB.analyzing import *
import os

#Restricted test problems
#-----------------------------------------------

TP_list = []

#TP1 - multiple global edge solutions
#-----------------------------------------------
def func(x): return x[0] + x[1]
def grad(x): return obvec([1,1])
def hess(x): return obmat([[0,0],[0,0]])

def cons_1(x): return -(x[0])**2 -(x[1])**2 +6.5
def cons_1_grad(x): return obvec([-2*(x[0]),-2*(x[1])])
def cons_1_hess(x): return obmat([[-2,0],[0,-2]])

def cons_2(x): return -x[0] +x[1] -2
def cons_2_grad(x): return obvec([-1,1])
def cons_2_hess(x): return obmat([[0,0],[0,0]])

def cons_3(x): return x[0] -x[1] -2
def cons_3_grad(x): return obvec([1,-1])
def cons_3_hess(x): return obmat([[0,0],[0,0]])

def cons_4(x): return (x[0])**2 +(x[1])**2 -16
def cons_4_grad(x): return obvec([2*(x[0]),2*(x[1])])
def cons_4_hess(x): return obmat([[2,0],[0,2]])

cons,cons_grad,cons_hess = [cons_1,cons_2,cons_3,cons_4],[cons_1_grad,cons_2_grad,cons_3_grad,cons_4_grad],[cons_1_hess,cons_2_hess,cons_3_hess,cons_4_hess]

X = intvec([[0,4.5],[0,4.5]])

TP_list.append(["TP1",(func,grad,hess,cons,cons_grad,cons_hess),X,(obvec([2.5,0.5]),obvec([0.5,2.5]))])

#TP2 - multiple global interior and edge solutions
#-----------------------------------------------
def func(x): return -0.05*((x[0]-2)**2+(x[1]-2)**2)**3 +0.8*((x[0]-2)**2+(x[1]-2)**2)
def grad(x): return obvec([-0.3*(x[0]-2)*((x[0]-2)**2+(x[1]-2)**2)**2 +1.6*(x[0]-2),-0.3*(x[1]-2)*((x[0]-2)**2+(x[1]-2)**2)**2 +1.6*(x[1]-2)])
def hess(x): return obmat([[-0.3*(4*((x[0]-2)**2 +(x[1]-2)**2)*(x[0]-2)**2 + ((x[0]-2)**2 +(x[1]-2)**2)**2) +1.6, -1.2*((x[0]-2)**2 +(x[1]-2)**2)*(x[1]-2)*(x[0]-2)],[-1.2*((x[0]-2)**2 +(x[1]-2)**2)*(x[0]-2)*(x[1]-2), -0.3*(4*((x[0]-2)**2 +(x[1]-2)**2)*(x[1]-2)**2 + ((x[0]-2)**2 +(x[1]-2)**2)) +1.6]])

def cons_1(x): return (x[0]-3)**3 -3 +x[1]
def cons_1_grad(x): return obvec([3*(x[0]-3)**2,1])
def cons_1_hess(x): return obmat([[6*(x[0]-3),0],[0,0]])

def cons_2(x): return -x[0] +x[1] -2
def cons_2_grad(x): return obvec([-1,1])
def cons_2_hess(x): return obmat([[0,0],[0,0]])

def cons_3(x): return x[0] -x[1] -2
def cons_3_grad(x): return obvec([1,-1])
def cons_3_hess(x): return obmat([[0,0],[0,0]])

def cons_4(x): return -log((x[0]+0.5)*(x[1]+0.5))+1
def cons_4_grad(x): return obvec([-1/(x[0]+0.5),-1/(x[1]+0.5)])
def cons_4_hess(x): return obmat([[1/(x[0]+0.5)**2,0],[0,1/(x[1]+0.5)**2]])

cons,cons_grad,cons_hess = [cons_1,cons_2,cons_3,cons_4],[cons_1_grad,cons_2_grad,cons_3_grad,cons_4_grad],[cons_1_hess,cons_2_hess,cons_3_hess,cons_4_hess]

X = intvec([[0,4.5],[0,4.5]])

TP_list.append(["TP2",(func,grad,hess,cons,cons_grad,cons_hess),X,(obvec([2,2]),obvec([4,2]),obvec([2,4]))])

#TP3 - multiple global solutions with disconnected feasible set
#-----------------------------------------------
def func(x): return (x[0]-2)**2+(x[1]-1)**2
def grad(x): return obvec([2*(x[0]-2),2*(x[1]-1)])
def hess(x): return obmat([[2,0],[0,2]])

def cons_1(x): return -(x[0]+1)**2 +x[1]
def cons_1_grad(x): return obvec([-2*(x[0]+1),1])
def cons_1_hess(x): return obmat([[-2,0],[0,0]])

def cons_2(x): return -(x[0]-2)**2 +x[1]
def cons_2_grad(x): return obvec([-2*(x[0]-2),1])
def cons_2_hess(x): return obmat([[-2,0],[0,0]])

def cons_3(x): return -(x[0]-5)**2 +x[1]
def cons_3_grad(x): return obvec([-2*(x[0]-5),1])
def cons_3_hess(x): return obmat([[-2,0],[0,0]])

def cons_4(x): return 1-x[1]
def cons_4_grad(x): return obvec([0,-1])
def cons_4_hess(x): return obmat([[0,0],[0,0]])

cons,cons_grad,cons_hess = [cons_1,cons_2,cons_3,cons_4],[cons_1_grad,cons_2_grad,cons_3_grad,cons_4_grad],[cons_1_hess,cons_2_hess,cons_3_hess,cons_4_hess]

X = intvec([[0,4.5],[0,4.5]])

TP_list.append(["TP3",(func,grad,hess,cons,cons_grad,cons_hess),X,(obvec([1,1]),obvec([3,1]))])

#TP4_1 - local minimum
#-----------------------------------------------
def func(x): return - 0.5*x[0]**2 - 0.5*x[1]**2
def grad(x): return obvec([-1*x[0],-1*x[1]])
def hess(x): return obmat([[-1,0],[0,-1]])

def cons_1(x): return (x[0]-2)**2 + (x[1]-1)**2 -4
def cons_1_grad(x): return obvec([2*(x[0]-2),2*(x[1]-1)])
def cons_1_hess(x): return obmat([[2,0],[0,2]])

def cons_MFB_2(x): return (-1/3)*(x[1]-4)**2 +x[0]
def cons_MFB_2_grad(x): return obvec([1,(-2/3)*(x[1]-4)])
def cons_MFB_2_hess(x): return obmat([[0,0],[0,-2/3]])

def cons_3(x): return -x[1] +1
def cons_3_grad(x): return obvec([0,-1])
def cons_3_hess(x): return obmat([[0,0],[0,0]])

cons,cons_grad,cons_hess = [cons_1,cons_MFB_2,cons_3],[cons_1_grad,cons_MFB_2_grad,cons_3_grad],[cons_1_hess,cons_MFB_2_hess,cons_3_hess]

X = intvec([[0,4.5],[0,4.5]])

TP_list.append(["TP4_1",(func,grad,hess,cons,cons_grad,cons_hess),X,(obvec([3,1]),)])

#TP4_2 - MFC violated
#-----------------------------------------------
def func(x): return - 0.5*x[0]**2 - 0.5*x[1]**2
def grad(x): return obvec([-1*x[0],-1*x[1]])
def hess(x): return obmat([[-1,0],[0,-1]])

def cons_1(x): return (x[0]-2)**2 + (x[1]-1)**2 -4
def cons_1_grad(x): return obvec([2*(x[0]-2),2*(x[1]-1)])
def cons_1_hess(x): return obmat([[2,0],[0,2]])

def cons_noMFB_2(x): return (1/9)*(x[0]-3)**3 -1 +x[1]
def cons_noMFB_2_grad(x): return obvec([(1/3)*(x[0]-3)**2,1])
def cons_noMFB_2_hess(x): return obmat([[(2/3)*(x[0]-3),0],[0,0]])

def cons_3(x): return -x[1] +1
def cons_3_grad(x): return obvec([0,-1])
def cons_3_hess(x): return obmat([[0,0],[0,0]])

cons,cons_grad,cons_hess = [cons_1,cons_noMFB_2,cons_3],[cons_1_grad,cons_noMFB_2_grad,cons_3_grad],[cons_1_hess,cons_noMFB_2_hess,cons_3_hess]

X = intvec([[0,4.5],[0,4.5]])

TP_list.append(["TP4_2",(func,grad,hess,cons,cons_grad,cons_hess),X,(obvec([3,1]),)])

#TP5 - spurious Points
#-----------------------------------------------
def func(x): return x[0] + x[1]
def grad(x): return obvec([1,1])
def hess(x): return obmat([[0,0],[0,0]])

def cons_1(x): return -(x[0]-1)**2 +x[1]-1
def cons_1_grad(x): return obvec([-2*(x[0]-1),1])
def cons_1_hess(x): return obmat([[-2,0],[0,0]])

def cons_2(x): return x[0] -x[1]
def cons_2_grad(x): return obvec([1,-1])
def cons_2_hess(x): return obmat([[0,0],[0,0]])

def cons_3(x): return  (x[0]-2)**2 +(x[1]-2)**2 -2
def cons_3_grad(x): return obvec([2*(x[0]-2),2*(x[1]-2)])
def cons_3_hess(x): return obmat([[2,0],[0,2]])

cons,cons_grad,cons_hess = [cons_1,cons_2,cons_3],[cons_1_grad,cons_2_grad,cons_3_grad],[cons_1_hess,cons_2_hess,cons_3_hess]

X = intvec([[0,4.5],[0,4.5]])

TP_list.append(["TP5",(func,grad,hess,cons,cons_grad,cons_hess),X,(obvec([1,1]),)])

#TP6_1 - infinite global solutions with disconnected solution set
#-----------------------------------------------
def func(x): return x[0] + x[1] -3
def grad(x): return obvec([1,1])
def hess(x): return obmat([[0,0],[0,0]])

def cons_1(x): return -(x[0]-2)**2 -(1/3)*(x[1]-1)**2 +1
def cons_1_grad(x): return obvec([-2*(x[0]-2),-(2/3)*(x[1]-1)])
def cons_1_hess(x): return obmat([[-2,0],[0,-2/3]])

def cons_2(x): return -x[0] -x[1] +4
def cons_2_grad(x): return obvec([-1,-1])
def cons_2_hess(x): return obmat([[0,0],[0,0]])

def cons_3(x): return -x[0] +0.5
def cons_3_grad(x): return obvec([-1,0])
def cons_3_hess(x): return obmat([[0,0],[0,0]])

def cons_4(x): return -x[1] +1
def cons_4_grad(x): return obvec([0,-1])
def cons_4_hess(x): return obmat([[0,0],[0,0]])

cons,cons_grad,cons_hess = [cons_1,cons_2,cons_3,cons_4],[cons_1_grad,cons_2_grad,cons_3_grad,cons_4_grad],[cons_1_hess,cons_2_hess,cons_3_hess,cons_4_hess]

X = intvec([[0,4.5],[0,4.5]])

TP_list.append(["TP6_1",(func,grad,hess,cons,cons_grad,cons_hess),X,(obvec([0.5,3.5]),obvec([1.5,2.5]),obvec([3,1]))])

#TP6_2 - infinite global solutions with connected solution set
#-----------------------------------------------
def func(x): return (x[0]-2)**2 +(1/3)*(x[1]-1)**2
def grad(x): return obvec([2*(x[0]-2),(2/3)*(x[1]-1)])
def hess(x): return obmat([[2,0],[0,2/3]])

def cons_1(x): return -(x[0]-2)**2 -(1/3)*(x[1]-1)**2 +1
def cons_1_grad(x): return obvec([-2*(x[0]-2),-(2/3)*(x[1]-1)])
def cons_1_hess(x): return obmat([[-2,0],[0,-2/3]])

def cons_2(x): return -x[0] -x[1] +4
def cons_2_grad(x): return obvec([-1,-1])
def cons_2_hess(x): return obmat([[0,0],[0,0]])

def cons_3(x): return -x[0] +0.5
def cons_3_grad(x): return obvec([-1,0])
def cons_3_hess(x): return obmat([[0,0],[0,0]])

def cons_4(x): return -x[1] +1
def cons_4_grad(x): return obvec([0,-1])
def cons_4_hess(x): return obmat([[0,0],[0,0]])

cons,cons_grad,cons_hess = [cons_1,cons_2,cons_3,cons_4],[cons_1_grad,cons_2_grad,cons_3_grad,cons_4_grad],[cons_1_hess,cons_2_hess,cons_3_hess,cons_4_hess]

X = intvec([[0,4.5],[0,4.5]])

TP_list.append(["TP6_2",(func,grad,hess,cons,cons_grad,cons_hess),X,(obvec([1.5,2.5]),obvec([2.5,2.5]),obvec([3,1]))])


#Parameter assignment
#-----------------------------------------------
epsilon_max_list = [0.5]*8
delta_max_list = [0.5]*8
bounding_procedure_list = [optimal_centered_forms]*8
max_time_list = [240]*8

fname = None #os.path.join(os.path.dirname(__file__), "../TP_probe")


#Code - Running the algorithm and collecting data
#-----------------------------------------------
probe_count = 0
for j,TP in enumerate(TP_list):
    TP_title,(func,grad,hess,cons,cons_grad,cons_hess),X,S = TP

    epsilon_max = epsilon_max_list[j] #if TP != "TP4_2" else 2
    delta_max = delta_max_list[j]
    bounding_procedure = bounding_procedure_list[j]
    max_time = max_time_list[j]

    #Code je TP
    #-----------------------------------------------
    #start = time.monotonic()
    solution, y_best ,k ,t ,W = analysed_impfunc_BandB(func, cons, X, bounding_procedure=bounding_procedure, grad=grad, hess=hess, cons_grad=cons_grad, cons_hess=cons_hess, epsilon=0, delta=0, epsilon_max=epsilon_max, delta_max=delta_max, max_time=max_time, save_lists=False)
    #total_time = time.monotonic() - start

    solution_Boxes = [Oi[0] for Oi in solution]
    
    edge_data = {"Tolerance epsilon_max": epsilon_max, "Tolerance delta_max": delta_max, "Bounding procedure": bounding_procedure.__name__, "Search": "Bf_search"}
    edge_data["Number of solution boxes"] = len(solution_Boxes)
    edge_data["Approximation includes S"] = all(any(all((si in Bi) for si, Bi in zip(s,B)) for B in solution_Boxes) for s in S)
    if y_best != None:
        y_best_str = "("
        for yi in y_best: y_best_str += f"{yi:g},"
        edge_data["Best incumbent"] = y_best_str.strip(",") + ")"
        edge_data["Best incumbent value"] = func(y_best)
    else:
        edge_data["Best incumbent"] = "Not found"
        edge_data["Best incumbent value"] = "-"
    edge_data["Iterations needed/done"] = k
    edge_data["Runtime [s]"] = t #total_time
    edge_data["Number of left boxes in W"] = len(W)
    
    try: 
        import pandas as pd
        from openpyxl import load_workbook
        df_TP_data = pd.Series(edge_data, name=f"{TP_title:8s}").to_frame().T
    except:
        if probe_count == 0: 
            print("Restricted test problems - Data:")
            print(f"{' ':8s}",f"{'Tolerance epsilon_max':>24s}",f"{'Tolerance delta_max':>24s}",f"{'Bounding procedure':>24s}",f"{'Search':>12s}",f"{'Number of solution boxes':>24s}",f"{'Approximation includes S':>24s}",f"{'Best incumbent':>24s}",f"{'Best incumbent value':>24s}",f"{'Iterations needed/done':>24s}",f"{'Runtime [s]':>12s}",f"{'Number of left boxes in W':>28s}")
        print(f"{TP_title:8s}",f"{edge_data['Tolerance epsilon_max']:>24.6f}",f"{edge_data['Tolerance delta_max']:>24.6f}",f"{edge_data['Bounding procedure']:>24s}",f"{edge_data['Search']:>12s}",f"{edge_data['Number of solution boxes']:>24d}",f"{edge_data['Approximation includes S']:>24b}",f"{edge_data['Best incumbent']:>24s}",f"{edge_data['Best incumbent value']:>24.6f}",f"{edge_data['Iterations needed/done']:>24d}",f"{edge_data['Runtime [s]']:>12.6f}",f"{edge_data['Number of left boxes in W']:>28d}")
    else:
        if probe_count == 0:
            print("Restricted test problems - Data:")
            print(df_TP_data.to_string(col_space=[24,24,24,12,24,24,24,24,24,12,28]))
        else:
            print(df_TP_data.to_string(header=False,col_space=[24,24,24,12,24,24,24,24,24,12,28]))
        if fname:
            try:
                wb = load_workbook(fname+".xlsx")
            except:
                with pd.ExcelWriter(fname+".xlsx", mode='w', engine='openpyxl') as writer:
                    df_TP_data.to_excel(writer, sheet_name="Probe_TP")
            else:
                ws = wb["Probe_TP"]
                ws["A"+str(probe_count+2)] = TP_title
                d = ws["A"+str(probe_count+2)]
                d.style = 'Pandas'
                for wert, col in zip(df_TP_data.values[0],range(2,13)):
                    ws.cell(row=probe_count+2, column=col, value=wert)
                wb.save(fname+".xlsx")
        
    probe_count += 1


#Boxrestricted test problems from Eichfelder
#-----------------------------------------------

BTP_list = []

#BTP1 - Easom
#-----------------------------------------------
def easom_func(x):
    return -cos(x[0])*cos(x[1])*exp(-(x[0] -np.pi)**2 -(x[1] -np.pi)**2)
def easom_grad(x):
    return obvec([(sin(x[0]) +2*x[0]*cos(x[0]) -2*np.pi*cos(x[0]))*cos(x[1])*exp(-(x[0]-np.pi)**2 -(x[1]-np.pi)**2),(sin(x[1]) +2*x[1]*cos(x[1]) -2*np.pi*cos(x[1]))*cos(x[0])*exp(-(x[0]-np.pi)**2 -(x[1]-np.pi)**2)])
def easom_hess(x):
    return obmat([[(4*np.pi*sin(x[0])-4*x[0]**2*cos(x[0])-(4*np.pi**2 -3)*cos(x[0])-x[0]*(4*sin(x[0]) -8*np.pi*cos(x[0])))*cos(x[1])*exp(-(x[0]-np.pi)**2 -(x[1]-np.pi)**2), (sin(x[0]) +2*x[0]*cos(x[0]) -2*np.pi*cos(x[0]))*(sin(x[1]) +2*x[1]*cos(x[1]) -2*np.pi*cos(x[1]))*exp(-(x[0]-np.pi)**2 -(x[1]-np.pi)**2)],
                  [(sin(x[0]) +2*x[0]*cos(x[0]) -2*np.pi*cos(x[0]))*(sin(x[1]) +2*x[1]*cos(x[1]) -2*np.pi*cos(x[1]))*exp(-(x[0]-np.pi)**2 -(x[1]-np.pi)**2), (4*np.pi*sin(x[1])-4*x[1]**2*cos(x[1])-(4*np.pi**2 -3)*cos(x[1])-x[1]*(4*sin(x[1]) -8*np.pi*cos(x[1])))*cos(x[0])*exp(-(x[0]-np.pi)**2 -(x[1]-np.pi)**2)]])

easom_X = intvec([[-100,100],[-100,100]])

BTP_list.append(["easom",(easom_func,easom_grad,easom_hess),easom_X,[]])

#BTP2 - Rastrigin
#-----------------------------------------------
def rastrigin_func(x):
    return 20 +x[0]**2 +x[1]**2 -10*(cos(2*np.pi*x[0]) +cos(2*np.pi*x[1]))
def rastrigin_grad(x):
    return obvec([2*x[0] +20*sin(2*np.pi*x[0])*np.pi, 2*x[1] +20*sin(2*np.pi*x[1])*np.pi])
def rastrigin_hess(x):
    return obmat([[2 +40*cos(2*np.pi*x[0])*np.pi**2, 0],[0, 2 +40*cos(2*np.pi*x[0])*np.pi**2]])

rastrigin_X = intvec([[-5.12,5.12],[-5.12,5.12]])

BTP_list.append(["rastrigin",(rastrigin_func,rastrigin_grad,rastrigin_hess),rastrigin_X,[]])

#BTP3 - Hump
#-----------------------------------------------
def hump_func(x):
    return (4 -2.1*x[0]**2 +(1/3)*x[0]**4)*x[0]**2 +x[0]*x[1] -(4 -4*x[1]**2)*x[1]**2
def hump_grad(x):
    return obvec([8*x[0] -8.4*x[0]**3 +2*x[0]**5 +x[1], x[0] -8*x[1] +16*x[1]**3])
def hump_hess(x):
    return obmat([[8 -25.2*x[0]**2 +10*x[0]**4, 1],[1, -8 +48*x[1]**2]])

hump_X = intvec([[-1.9,1.9],[-1.1,1.1]])

BTP_list.append(["hump",(hump_func,hump_grad,hump_hess),hump_X,[]])

#BTP4 - Branin
#-----------------------------------------------
def branin_func(x):
    return (x[1] -(5.1/(4*np.pi**2))*x[0]**2 +(5/np.pi)*x[0] -6)**2 +10*(1-(1/(8*np.pi)))*cos(x[0]) +10
def branin_grad(x):
    return obvec([2*(x[1] -(5.1/(4*np.pi**2))*x[0]**2 +(5/np.pi)*x[0] -6)*(-(5.1/(2*np.pi**2))*x[0] +(5/np.pi)) -10*(1-(1/(8*np.pi)))*sin(x[0]), 2*(x[1] -(5.1/(4*np.pi**2))*x[0]**2 +(5/np.pi)*x[0] -6)])
def branin_hess(x):
    return obmat([[(-5/np.pi**2)*x[1] +(78.03/(4*np.pi**4))*x[0]**2 -(76.5/np.pi**3)*x[0] +(80.6/np.pi**2) -10*(1-(1/(8*np.pi)))*cos(x[0]), (-5.1/np.pi**2)*x[0] +(10/np.pi)],[(-5.1/np.pi**2)*x[0] +(10/np.pi), 2]])

branin_X = intvec([[-5,10],[0,15]])

BTP_list.append(["branin",(branin_func,branin_grad,branin_hess),branin_X,[]])

#BTP5 - Himmelblau
#-----------------------------------------------
def himmelblau_func(x):
    return (x[0]**2 +x[1] -11)**2 + (x[0] +x[1]**2 -7)**2
def himmelblau_grad(x):
    return obvec([4*(x[0]**2 +x[1] -11)*x[0] +2*(x[0] +x[1]**2 -7), 2*(x[0]**2 +x[1] -11) +4*(x[0] +x[1]**2 -7)*x[1]])
def himmelblau_hess(x):
    return obmat([[12*x[0]**2 +4*x[1] -42, 4*x[0] +4*x[1]],[4*x[0] +4*x[1], 4*x[0] +12*x[1]**2 -26]])

himmelblau_X = intvec([[-6,6],[-6,6]])

BTP_list.append(["himmelblau",(himmelblau_func,himmelblau_grad,himmelblau_hess),himmelblau_X,[]])

#BTP6 - Rastrigin modification
#-----------------------------------------------
def rastrigin_mod_func(x):
    return 20 +x[0]**2 +x[1]**2 +10*(cos(2*np.pi*x[0]) +cos(2*np.pi*x[1]))
def rastrigin_mod_grad(x):
    return obvec([2*x[0] -20*sin(2*np.pi*x[0])*np.pi, 2*x[1] -20*sin(2*np.pi*x[1])*np.pi])
def rastrigin_mod_hess(x):
    return obmat([[2 -40*cos(2*np.pi*x[0])*np.pi**2, 0],[0, 2 -40*cos(2*np.pi*x[0])*np.pi**2]])

rastrigin_mod_X = intvec([[-5.12,5.12],[-5.12,5.12]])

BTP_list.append(["rastrigin_mod",(rastrigin_mod_func,rastrigin_mod_grad,rastrigin_mod_hess),rastrigin_mod_X,[]])

#BTP7 - Shubert
#-----------------------------------------------
def shubert_func(x):
    return sum(i *cos((i +1)*x[0] +i) for i in range(1,6)) *sum(j *cos((j +1)*x[1] +j) for j in range(1,6))
def shubert_grad(x):
    return obvec([sum(-i *sin((i +1)*x[0] +i) *(i +1) for i in range(1,6)) *sum(j *cos((j +1)*x[1] +j) for j in range(1,6)), sum(-j *sin((j +1)*x[1] +j) *(j +1) for j in range(1,6)) *sum(i *cos((i +1)*x[0] +i) for i in range(1,6))])
def shubert_hess(x):
    return obmat([[sum(-i *cos((i +1)*x[0] +i) *(i +1)**2 for i in range(1,6)) *sum(j *cos((j +1)*x[1] +j) for j in range(1,6)), sum(-i *sin((i +1)*x[0] +i) *(i +1) for i in range(1,6)) *sum(-j *sin((j +1)*x[1] +j) *(j +1) for j in range(1,6))],
                  [sum(-i *sin((i +1)*x[0] +i) *(i +1) for i in range(1,6)) *sum(-j *sin((j +1)*x[1] +j) *(j +1) for j in range(1,6)), sum(-j *sin((j +1)*x[1] +j) *(j +1)**2 for j in range(1,6)) *sum(i *cos((i +1)*x[0] +i) for i in range(1,6))]])

shubert_X = intvec([[-10,10],[-10,10]])

BTP_list.append(["shubert",(shubert_func,shubert_grad,shubert_hess),shubert_X,[]])

#BTP8 - Deb1
#-----------------------------------------------
def sinN(n,x):
    res = sin(x)
    for i in range(1,n):
        res = sin(res)
    return res
def cosN(n,x):
    res = cos(x)
    for i in range(1,n):
        res = cos(res)
    return res
def deb1_func(x):
    return -0.5*(sinN(6,5*np.pi*x[0]) + sinN(6,5*np.pi*x[1]))
def deb1_grad(x):
    return obvec([-47.1239 *sinN(5,5*np.pi*x[0])*cos(5*np.pi*x[0]), -47.1239 *sinN(5,5*np.pi*x[1])*cos(5*np.pi*x[1])])
def deb1_hess(x):
    return obmat([[740.22*sinN(6,5*np.pi*x[0]) -3701.1*sinN(4,5*np.pi*x[0])*cosN(2,5*np.pi*x[0]),0],[0,740.22*sinN(6,5*np.pi*x[1]) -3701.1*sinN(4,5*np.pi*x[1])*cosN(2,5*np.pi*x[1])]])

deb1_X = intvec([[0,1],[0,1]])

BTP_list.append(["deb1",(deb1_func,deb1_grad,deb1_hess),deb1_X,[]])

#BTP9 - Vincent
#-----------------------------------------------
def vincent_func(x):
    return -0.5*(sin(10*log(x[0])) +sin(10*log(x[1])))
def vincent_grad(x):
    return obvec([-(5/x[0])*cos(10*log(x[0])), -(5/x[1])*cos(10*log(x[1]))])
def vincent_hess(x):
    return obmat([[(5/x[0]**2)*(10*sin(10*log(x[0])) +cos(10*log(x[0]))), 0],[0, (5/x[1]**2)*(10*sin(10*log(x[1])) +cos(10*log(x[1])))]])

vincent_X = intvec([[0.25,10],[0.25,10]])

BTP_list.append(["vincent",(vincent_func,vincent_grad,vincent_hess),vincent_X,[]])


#Parameter assignment
#-----------------------------------------------
epsilon_max_list = [0.5]*9
bounding_procedure_list = [optimal_centered_forms]*9
max_time_list = [240]*9

fname = None #os.path.join(os.path.dirname(__file__), "../BTP_probe")


#Code - Running the algorithm and collecting data
#-----------------------------------------------
probe_count = 0
for j,BTP in enumerate(BTP_list):
    BTP_title,(func,grad,hess),X,S = BTP

    epsilon_max = epsilon_max_list[j]
    bounding_procedure = bounding_procedure_list[j]
    max_time = max_time_list[j]

    #Code je BTP
    #-----------------------------------------------
    #start = time.monotonic()
    solution, y_best ,k ,t ,W = analysed_impfunc_boxres_BandB(func, X, bounding_procedure=bounding_procedure, grad=grad, hess=hess, epsilon=0, epsilon_max=epsilon_max, max_time=max_time, save_lists=False)
    #total_time = time.monotonic() - start

    solution_Boxes = [Oi[0] for Oi in solution]
    
    edge_data = {"Tolerance epsilon_max": epsilon_max, "Bounding procedure": bounding_procedure.__name__, "Search": "Bf_search"}
    edge_data["Number of solution boxes"] = len(solution_Boxes)
    edge_data["Approximation includes S"] = all(any(all((si in Bi) for si, Bi in zip(s,B)) for B in solution_Boxes) for s in S)
    if y_best != None:
        y_best_str = "("
        for yi in y_best: y_best_str += f"{yi:g},"
        edge_data["Best incumbent"] = y_best_str.strip(",") + ")"
        edge_data["Best incumbent value"] = func(y_best)
    else:
        edge_data["Best incumbent"] = "Not found"
        edge_data["Best incumbent value"] = "-"
    edge_data["Iterations needed/done"] = k
    edge_data["Runtime [s]"] = t #total_time
    edge_data["Number of left boxes in W"] = len(W)
    
    try: 
        import pandas as pd
        from openpyxl import load_workbook
        df_BTP_data = pd.Series(edge_data, name=f"{BTP_title:14s}").to_frame().T
    except:
        if probe_count == 0: 
            print("Boxrestricted test problems - Data:")
            print(f"{' ':14s}",f"{'Tolerance epsilon_max':>24s}",f"{'Bounding procedure':>24s}",f"{'Search':>12s}",f"{'Number of solution boxes':>24s}",f"{'Approximation includes S':>24s}",f"{'Best incumbent':>24s}",f"{'Best incumbent value':>24s}",f"{'Iterations needed/done':>24s}",f"{'Runtime [s]':>12s}",f"{'Number of left boxes in W':>28s}")
        print(f"{BTP_title:14s}",f"{edge_data['Tolerance epsilon_max']:>24.6f}",f"{edge_data['Bounding procedure']:>24s}",f"{edge_data['Search']:>12s}",f"{edge_data['Number of solution boxes']:>24d}",f"{edge_data['Approximation includes S']:>24b}",f"{edge_data['Best incumbent']:>24s}",f"{edge_data['Best incumbent value']:>24.6f}",f"{edge_data['Iterations needed/done']:>24d}",f"{edge_data['Runtime [s]']:>12.6f}",f"{edge_data['Number of left boxes in W']:>28d}")
    else:
        if probe_count == 0:
            print("Boxrestricted test problems - Data:")
            print(df_BTP_data.to_string(col_space=[24,24,12,24,24,24,24,24,12,28]))
        else:
            print(df_BTP_data.to_string(header=False,col_space=[24,24,12,24,24,24,24,24,12,28]))
        if fname:
            try:
                wb = load_workbook(fname+".xlsx")
            except:
                with pd.ExcelWriter(fname+".xlsx", mode='w', engine='openpyxl') as writer:
                    df_BTP_data.to_excel(writer, sheet_name="Probe_TP")
            else:
                ws = wb["Probe_TP"]
                ws["A"+str(probe_count+2)] = BTP_title
                d = ws["A"+str(probe_count+2)]
                d.style = 'Pandas'
                for wert, col in zip(df_BTP_data.values[0],range(2,12)):
                    ws.cell(row=probe_count+2, column=col, value=wert)
                wb.save(fname+".xlsx")
        
    probe_count += 1